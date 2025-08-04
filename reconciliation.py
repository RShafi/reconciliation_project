import streamlit as st
import pandas as pd
import re
from io import BytesIO
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from datetime import datetime, timedelta
import os

st.set_page_config(page_title="ACH Invoice Processor", layout="centered")

st.title("ACH Invoice Processor (by Payment Date)")
st.write("Upload your Supplier Invoice Excel file and enter ACH information below.")

# --- User Inputs ---
ach_amount = st.text_input("ACH Amount ($)")
ach_description = st.text_input("ACH Description (optional)")
ach_date = st.date_input("ACH Payment Date")
uploaded_file = st.file_uploader("Upload Supplier Invoice Excel File", type=["xlsx"])

# --- Extract Consultant Info from Line Item ---
def parse_line_item(description):
    match = re.match(r"(.+?) \((S\d+)\)\[C\]:(\d{4}-\d{2}-\d{2}):(\d{4}-\d{2}-\d{2})", str(description))
    if match:
        full_name = match.group(1).strip().split()
        return pd.Series([
            full_name[0],
            " ".join(full_name[1:]) if len(full_name) > 1 else "",
            match.group(2),
            match.group(4)
        ])
    return pd.Series([None, None, None, None])

# --- Generate Excel Output ---
if st.button("Generate Excel Output"):
    if not ach_amount or not ach_description or not ach_date or not uploaded_file:
        st.error("❌ Please fill in all fields and upload a file.")
    else:
        try:
            ach_amount_float = float(ach_amount)
            df = pd.read_excel(uploaded_file)
            
            df["Payment Date"] = pd.to_datetime(df["Payment Date"], errors='coerce').dt.date
            ach_date_only = ach_date.strftime('%Y-%m-%d')
            matching_df = df[df["Payment Date"] == datetime.strptime(ach_date_only, '%Y-%m-%d').date()]


            if matching_df.empty:
                st.error("❌ No rows found with Payment Date matching ACH Date.")
            else:
                total = matching_df["Extended Amount"].sum()
                if round(total, 2) != round(ach_amount_float, 2):
                    st.error(f"❌ ACH amount mismatch!\nExpected total from file: ${total:.2f}\nEntered: ${ach_amount_float:.2f}")
                else:
                    # Extract consultant info
                    matching_df[['First Name', 'Last Name', 'Candidate CAI ID', 'Vector Week Ending']] = matching_df[
                        'Line Item Description'].apply(parse_line_item)

                    # Add FMS Week Ending (Vector + 2 days)
                    matching_df["FMS Week Ending"] = pd.to_datetime(matching_df["Vector Week Ending"], errors='coerce') + timedelta(days=2)
                    matching_df["FMS Week Ending"] = matching_df["FMS Week Ending"].dt.strftime('%Y-%m-%d')

                    # Add ACH metadata
                    matching_df['ACH Description'] = ach_description
                    matching_df['ACH Amount'] = ach_amount_float
                    matching_df['ACH Date'] = ach_date.strftime('%Y-%m-%d')

                    # Rearrange output columns
                    df_final = matching_df[[
                        'First Name', 'Last Name', 'Candidate CAI ID', 'Vector Week Ending', 'FMS Week Ending',
                        'Quantity', 'Unit Cost', 'Extended Amount', 'Invoice Amount', 'Due Date',
                        'CAI Invoice Number', "Supplier's Invoice Number",
                        'ACH Description', 'ACH Amount', 'ACH Date'
                    ]].rename(columns={
                        'Quantity': 'Hours',
                        'Unit Cost': 'Bill Rate',
                        "Supplier's Invoice Number": 'ESG Invoice Number'
                    })

                    # Sort by Vector Week Ending
                    df_final = df_final.sort_values(by="Vector Week Ending")

                    # Format output with summary
                    summary_fields = ['ACH Description', 'ACH Amount', 'ACH Date']
                    summary_data = {field: df_final[field].iloc[0] for field in summary_fields}
                    df_final = df_final.drop(columns=summary_fields)

                    # Insert blank rows between week endings
                    formatted_rows = []
                    prev_week = None
                    for _, row in df_final.iterrows():
                        current_week = row["Vector Week Ending"]
                        if prev_week and current_week != prev_week:
                            formatted_rows.append([""] * len(df_final.columns))  # blank row
                        formatted_rows.append(row.tolist())
                        prev_week = current_week

                    # Write to Excel
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "ACH Report"

                    # Add summary
                    ws.append(["ACH Summary"])
                    for key, value in summary_data.items():
                        ws.append([key, value])
                    ws.append([])  # blank line

                    # Add headers
                    ws.append(df_final.columns.tolist())
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)

                    # Add rows
                    for row in formatted_rows:
                        ws.append(row)

                    # Auto-size columns
                    for col in ws.columns:
                        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

                    # Export
                    buffer = BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)

                    st.success("✅ ACH matched. Download your formatted report below.")
                    st.download_button("Download Excel Report", data=buffer, file_name="ACH_By_Date_Report.xlsx")

        except ValueError:
            st.error("❌ ACH Amount must be numeric.")
        except Exception as e:
            st.error(f"❌ Unexpected error: {e}")

st.markdown("---")  # horizontal divider
if st.button("❌ Exit Program"):
    st.warning("Shutting down the Streamlit app...")
    os._exit(0)